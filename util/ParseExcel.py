#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@author: kyle
@time: 2018/3/7 10:00
"""
import openpyxl
from openpyxl.styles import Font,Border,Side
import time


class ParseExcel(object):

    def __init__(self):
        self.workbook = None
        self.excelFile = None
        # 设置字体颜色
        self.font = Font(color=None)
        # 颜色值对应RGB值
        self.RGBDict = {'red': 'FFFF3030', 'green': 'FF008B00'}

    def loadWorkBook(self, excelPathAndName):
        # 将excel文件加载到内存中，获取workbook对象
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        else:
            self.excelFile = excelPathAndName
            return self.workbook

    def getSheetByName(self, sheetName):
        # 根据sheet名获取对应sheet对象
        try:
            sheet = self.workbook[sheetName]
        except Exception as e:
            raise e
        else:
            return sheet

    def getSheetByIndex(self, sheetIndex):
        # 根据索引获取对应sheet对象
        try:
            sheetname = self.workbook.sheetnames[sheetIndex]
        except Exception as e:
            raise e
        sheet = self.workbook[sheetname]
        return sheet

    def getRowNumber(self, sheet):
        # 获取sheet中有数据区域的行数
        return sheet.max_row

    def getColNumber(self, sheet):
        # 获取sheet中有数据区域的列数
        return sheet.max_column

    def getStartRowNumber(self, sheet):
        # 获取sheet中有数据区域的开始行号
        return sheet.min_row

    def getStartColNumber(self, sheet):
        # 获取sheet中有数据区域的开始列号
        return sheet.min_column

    def getRow(self, sheet, rowNo):
        # 获取一行数据
        try:
            return list(sheet.rows)[rowNo - 1]
        except Exception as e:
            raise e

    def getColumn(self, sheet, colNo):
        # 获取一列数据
        try:
            return list(sheet.column)[colNo - 1]
        except Exception as e:
            raise e

    def getCellOfValue(self, sheet, coordinate=None, rowNo=None, colNo=None):
        # 根据单元格所在位置获取单元格的值
        if coordinate != None:
            try:
                return sheet[coordinate].value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinate of cell!")

    def getCellOfObject(self, sheet, coordinate=None, rowNo=None, colNo=None):
        # 获取单元格对象
        if coordinate != None:
            try:
                return sheet[coordinate]
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinate of Cell!")

    def writeCell(self, sheet, content, coordinate=None, rowNo=None, colNo=None, style=None):
        # 根据单元格在excel中编码坐标或者数字索引向单元格中写入数据
        if coordinate != None:
            try:
                sheet[coordinate].value = content
                if style is not None:
                    sheet[coordinate].font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo, column=colNo).value = content
                if style is not None:
                    sheet.cell(row=rowNo, column=colNo).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinate of cell!")

    def writeCellCurrentTime(self,sheet, coordinate=None, rowNo=None, colNo=None):
        # 将当前时间写入单元格
        now = int(time.time())
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
        if coordinate != None:
            try:
                sheet[coordinate].value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo, column=colNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinate of cell!")
